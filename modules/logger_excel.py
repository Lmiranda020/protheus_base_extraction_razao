import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOG_FOLDER   = "log"
LOG_FILENAME = "log_automacao.xlsx"

# Paleta de cores
COR_HEADER      = "1F4E79"   # azul escuro
COR_SUCESSO     = "E2EFDA"   # verde claro
COR_ERRO        = "FDDCDC"   # vermelho claro
COR_RESUMO_OK   = "70AD47"   # verde
COR_RESUMO_ERR  = "FF0000"   # vermelho
COR_LINHA_PAR   = "F2F7FF"   # azul muito claro (zebra)

BORDA_FINA = Border(
    left   = Side(style="thin"),
    right  = Side(style="thin"),
    top    = Side(style="thin"),
    bottom = Side(style="thin"),
)

COLUNAS = [
    ("Data/Hora Início",    20),
    ("Data/Hora Fim",       20),
    ("Tipo",                16),
    ("Competência",         16),
    ("Filial",              12),
    ("Status",              12),
    ("Mensagem",            50),
    ("Tempo (s)",           12),
]


def _caminho_log(raiz_projeto: str) -> str:
    """Retorna o caminho completo do arquivo de log dentro da pasta log/."""
    pasta_log = os.path.join(raiz_projeto, LOG_FOLDER)
    os.makedirs(pasta_log, exist_ok=True)   # cria a pasta se não existir
    return os.path.join(pasta_log, LOG_FILENAME)


def _criar_planilha(wb: Workbook) -> None:
    """Cria a aba 'Log' com cabeçalho formatado"""
    ws = wb.active
    ws.title = "Log"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color=COR_HEADER)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (titulo, largura) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = BORDA_FINA
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _estilizar_linha(ws, row: int, status: str) -> None:
    is_par = (row % 2 == 0)
    bg_cor = COR_SUCESSO if status == "✅ Sucesso" else (
             COR_ERRO    if status == "❌ Erro"    else
             (COR_LINHA_PAR if is_par else "FFFFFF"))

    fill = PatternFill("solid", start_color=bg_cor)
    for col in range(1, len(COLUNAS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = fill
        cell.border    = BORDA_FINA
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font      = Font(name="Arial", size=10)

    # Coluna Status — negrito e colorido
    status_cell = ws.cell(row=row, column=6)
    cor_txt = COR_RESUMO_OK if status == "✅ Sucesso" else (
              COR_RESUMO_ERR if status == "❌ Erro" else "000000")
    status_cell.font = Font(name="Arial", size=10, bold=True, color=cor_txt)


class LogExecucao:
    """
    Gerencia o log incremental de execuções em Excel.
    O arquivo é salvo em: <raiz_projeto>/log/log_automacao.xlsx

    Uso:
        log = LogExecucao(raiz_projeto="C:/RPA")
        log.iniciar_execucao(tipo="Razão", competencia="30/04/2025",
                             filiais=["0101", "0102"])
        log.registrar_filial("0101", sucesso=True)
        log.registrar_filial("0102", sucesso=False, mensagem="Timeout ao exportar")
        resumo = log.finalizar_execucao()
    """

    def __init__(self, raiz_projeto: str):
        self.caminho     = _caminho_log(raiz_projeto)
        self.tipo        = ""
        self.competencia = ""
        self.filiais     = []
        self.inicio      = None
        self._registros  = []   # buffer: tuplas com os dados de cada filial


    def iniciar_execucao(self, tipo: str, competencia: str, filiais: list) -> None:
        """Marca o início da execução e guarda os metadados."""
        self.tipo        = tipo
        self.competencia = competencia
        self.filiais     = filiais
        self.inicio      = datetime.now()
        self._registros  = []
        print(f"📋 Log iniciado — Tipo: {tipo} | Competência: {competencia} "
              f"| Filiais: {len(filiais)}")
        print(f"📂 Arquivo de log: {self.caminho}")

    def registrar_filial(
        self,
        filial: str,
        sucesso: bool,
        mensagem: str = "",
        inicio_filial: datetime = None,
    ) -> None:
        """Registra o resultado de uma filial no buffer."""
        fim      = datetime.now()
        inicio_f = inicio_filial or fim
        tempo    = round((fim - inicio_f).total_seconds(), 1)
        status   = "✅ Sucesso" if sucesso else "❌ Erro"

        self._registros.append((
            inicio_f.strftime("%d/%m/%Y %H:%M:%S"),
            fim.strftime("%d/%m/%Y %H:%M:%S"),
            self.tipo,
            self.competencia,
            filial,
            status,
            mensagem,
            tempo,
        ))
        print(f"  {'✅' if sucesso else '❌'} Filial {filial} registrada no log "
              f"({tempo}s) — {status}")

    def finalizar_execucao(self) -> dict:
        """
        Persiste todos os registros no Excel e retorna um dicionário-resumo
        para ser usado no e-mail.
        """
        fim_geral = datetime.now()
        self._salvar_no_excel()

        sucesso_lst = [r for r in self._registros if r[5] == "✅ Sucesso"]
        erro_lst    = [r for r in self._registros if r[5] == "❌ Erro"]

        tempo_total = round((fim_geral - self.inicio).total_seconds(), 1) if self.inicio else 0

        resumo = {
            "tipo"        : self.tipo,
            "competencia" : self.competencia,
            "inicio"      : self.inicio.strftime("%d/%m/%Y %H:%M:%S") if self.inicio else "",
            "fim"         : fim_geral.strftime("%d/%m/%Y %H:%M:%S"),
            "tempo_total" : tempo_total,
            "total"       : len(self._registros),
            "sucesso"     : len(sucesso_lst),
            "erros"       : len(erro_lst),
            "filiais_ok"  : [r[4] for r in sucesso_lst],
            "filiais_err" : [r[4] for r in erro_lst],
            "caminho_log" : self.caminho,
        }

        print(f"\n📊 Resumo da execução — {self.tipo}")
        print(f"   Total : {resumo['total']} filiais")
        print(f"   ✅ OK  : {resumo['sucesso']}")
        print(f"   ❌ Err : {resumo['erros']}")
        print(f"   ⏱️  Tempo total: {tempo_total}s")

        return resumo


    def _salvar_no_excel(self) -> None:
        if os.path.exists(self.caminho):
            wb = load_workbook(self.caminho)
            ws = wb["Log"]
        else:
            wb = Workbook()
            _criar_planilha(wb)
            ws = wb["Log"]

        prox_linha = ws.max_row + 1

        for registro in self._registros:
            for col_idx, valor in enumerate(registro, start=1):
                ws.cell(row=prox_linha, column=col_idx, value=valor)
            _estilizar_linha(ws, prox_linha, registro[5])
            prox_linha += 1

        # Auto-filtro cobre toda a tabela
        ultima_col = get_column_letter(len(COLUNAS))
        ws.auto_filter.ref = f"A1:{ultima_col}{ws.max_row}"

        try:
            wb.save(self.caminho)
            print(f"💾 Log salvo em: {self.caminho}")
        except PermissionError:
            alt = self.caminho.replace(".xlsx", f"_{datetime.now().strftime('%H%M%S')}.xlsx")
            wb.save(alt)
            print(f"⚠️  Arquivo em uso. Log salvo em: {alt}")
            self.caminho = alt